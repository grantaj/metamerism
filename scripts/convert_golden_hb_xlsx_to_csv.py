#!/usr/bin/env python3
"""Convert the GOLDEN Heavy Body Acrylic spectral workbook to clean CSV files.

Input workbook expected:
    Reflectance Data for Golden HB 10 mil Drawdowns over White.xlsx

Outputs:
    golden_hb_paints.csv
    golden_hb_reflectance_percent_wide.csv
    golden_hb_reflectance_fraction_wide.csv
    golden_hb_ks_wide.csv
    golden_hb_spectra_long.csv
    README_golden_clean_csv.md

The original workbook has a visual two-row header with merged cells:
    - C:E     D65/10° Lab values
    - G:AK    % Reflectance values, 400--700 nm
    - AM:BQ   K/S values, 400--700 nm

CSV cannot represent those merged header groups, so this script writes explicit,
flat, reproducible tables that are convenient to load from Python.
"""

from __future__ import annotations

import argparse
import csv
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

EXPECTED_WAVELENGTHS = list(range(400, 701, 10))


@dataclass(frozen=True)
class PaintSpectrum:
    """One paint row from the GOLDEN workbook."""

    product_id: int | str
    paint_name: str
    lab_l_d65_10: float
    lab_a_d65_10: float
    lab_b_d65_10: float
    reflectance_percent: list[float]
    ks: list[float]


def _as_float(value: object, *, row: int, column: int) -> float:
    """Convert a workbook cell value to float with a useful error message."""
    if value is None:
        raise ValueError(f"Missing numeric value at row {row}, column {column}")
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Expected numeric value at row {row}, column {column}, got {value!r}"
        ) from exc


def _normalise_name(value: object) -> str:
    """Normalise paint names while preserving the workbook's wording."""
    if value is None:
        return ""
    return str(value).strip()


def _validate_header(ws: Worksheet) -> None:
    """Check that the workbook matches the expected GOLDEN sheet layout."""
    checks = {
        "C1": "D65/10°",
        "G1": "% Reflectance",
        "AM1": "K/S",
        "A2": "Prod #",
        "B2": "Name",
        "C2": "L*",
        "D2": "a*",
        "E2": "b*",
    }
    for cell, expected in checks.items():
        actual = ws[cell].value
        if actual != expected:
            raise ValueError(
                f"Unexpected workbook layout: cell {cell} should be {expected!r}, "
                f"but is {actual!r}"
            )

    reflectance_wavelengths = [ws.cell(2, col).value for col in range(7, 38)]
    ks_wavelengths = [ws.cell(2, col).value for col in range(39, 70)]

    if reflectance_wavelengths != EXPECTED_WAVELENGTHS:
        raise ValueError(
            "Unexpected reflectance wavelength header. "
            f"Expected {EXPECTED_WAVELENGTHS}, got {reflectance_wavelengths}"
        )
    if ks_wavelengths != EXPECTED_WAVELENGTHS:
        raise ValueError(
            "Unexpected K/S wavelength header. "
            f"Expected {EXPECTED_WAVELENGTHS}, got {ks_wavelengths}"
        )


def read_golden_workbook(
    xlsx_path: Path,
    sheet_name: str | None = None,
) -> list[PaintSpectrum]:
    """Read the GOLDEN workbook and return paint spectra."""
    workbook = load_workbook(xlsx_path, data_only=True, read_only=False)
    if sheet_name is None:
        sheet_name = workbook.sheetnames[0]
    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Sheet {sheet_name!r} not found. Available sheets: "
            f"{workbook.sheetnames}"
        )

    ws = workbook[sheet_name]
    _validate_header(ws)

    paints: list[PaintSpectrum] = []
    for row in range(3, ws.max_row + 1):
        product_id = ws.cell(row, 1).value
        paint_name = _normalise_name(ws.cell(row, 2).value)

        # Stop/skip completely empty trailing rows.
        if product_id is None and not paint_name:
            continue

        lab_l = _as_float(ws.cell(row, 3).value, row=row, column=3)
        lab_a = _as_float(ws.cell(row, 4).value, row=row, column=4)
        lab_b = _as_float(ws.cell(row, 5).value, row=row, column=5)

        reflectance_percent = [
            _as_float(ws.cell(row, col).value, row=row, column=col)
            for col in range(7, 38)
        ]
        ks = [
            _as_float(ws.cell(row, col).value, row=row, column=col)
            for col in range(39, 70)
        ]

        paints.append(
            PaintSpectrum(
                product_id=product_id,
                paint_name=paint_name,
                lab_l_d65_10=lab_l,
                lab_a_d65_10=lab_a,
                lab_b_d65_10=lab_b,
                reflectance_percent=reflectance_percent,
                ks=ks,
            )
        )

    if not paints:
        raise ValueError("No paint rows found in workbook")

    return paints


def _write_csv(
    path: Path,
    header: Sequence[str],
    rows: Iterable[Sequence[object]],
) -> None:
    """Write a CSV with stable newline and UTF-8 behaviour."""
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)


def write_clean_csvs(paints: Sequence[PaintSpectrum], out_dir: Path) -> None:
    """Write the cleaned GOLDEN CSV files."""
    out_dir.mkdir(parents=True, exist_ok=True)

    wavelength_columns = [f"wl_{w}" for w in EXPECTED_WAVELENGTHS]
    base_columns = ["product_id", "paint_name"]

    _write_csv(
        out_dir / "golden_hb_paints.csv",
        [
            "product_id",
            "paint_name",
            "lab_l_d65_10",
            "lab_a_d65_10",
            "lab_b_d65_10",
            "lab_illuminant",
            "lab_observer",
        ],
        (
            [
                p.product_id,
                p.paint_name,
                p.lab_l_d65_10,
                p.lab_a_d65_10,
                p.lab_b_d65_10,
                "D65",
                "10 degree",
            ]
            for p in paints
        ),
    )

    _write_csv(
        out_dir / "golden_hb_reflectance_percent_wide.csv",
        [*base_columns, *wavelength_columns],
        ([p.product_id, p.paint_name, *p.reflectance_percent] for p in paints),
    )

    _write_csv(
        out_dir / "golden_hb_reflectance_fraction_wide.csv",
        [*base_columns, *wavelength_columns],
        (
            [
                p.product_id,
                p.paint_name,
                *[value / 100.0 for value in p.reflectance_percent],
            ]
            for p in paints
        ),
    )

    _write_csv(
        out_dir / "golden_hb_ks_wide.csv",
        [*base_columns, *wavelength_columns],
        ([p.product_id, p.paint_name, *p.ks] for p in paints),
    )

    long_rows = []
    for p in paints:
        for wavelength, reflectance_percent, ks in zip(
            EXPECTED_WAVELENGTHS, p.reflectance_percent, p.ks, strict=True
        ):
            long_rows.append(
                [
                    p.product_id,
                    p.paint_name,
                    wavelength,
                    reflectance_percent,
                    reflectance_percent / 100.0,
                    ks,
                ]
            )

    _write_csv(
        out_dir / "golden_hb_spectra_long.csv",
        [
            "product_id",
            "paint_name",
            "wavelength_nm",
            "reflectance_percent",
            "reflectance_fraction",
            "ks",
        ],
        long_rows,
    )


def write_readme(out_dir: Path, paint_count: int) -> None:
    """Write a concise README describing the generated CSV files."""
    sample_count = len(EXPECTED_WAVELENGTHS)
    long_rows = paint_count * sample_count
    readme = f"""# Clean CSV export: GOLDEN Heavy Body spectral data

This directory contains a cleaned CSV export of the GOLDEN Heavy Body Acrylic
spectral data workbook:

`Reflectance Data for Golden HB 10 mil Drawdowns over White.xlsx`

The original Excel workbook uses a two-row header with merged cells. That is
convenient visually, but awkward in CSV because CSV has no way to represent
merged header groups. A direct conversion therefore produces confusing columns
such as `Unnamed: 7`, `Unnamed: 8`, and duplicate wavelength labels. These files
flatten the workbook into regular, Python-friendly tables.

## Files

| File | Purpose |
|---|---|
| `golden_hb_paints.csv` | Paint identity and CIE Lab values. |
| `golden_hb_reflectance_percent_wide.csv` | Reflectance as percent values. |
| `golden_hb_reflectance_fraction_wide.csv` | Reflectance scaled to `0-1`. |
| `golden_hb_ks_wide.csv` | One row per paint, with Kubelka-Munk `K/S` values. |
| `golden_hb_spectra_long.csv` | Tidy reflectance and `K/S` values. |

## Dataset shape

- Paints: **{paint_count}**
- Wavelength range: **{EXPECTED_WAVELENGTHS[0]}-{EXPECTED_WAVELENGTHS[-1]} nm**
- Wavelength interval: **10 nm**
- Samples per paint: **{sample_count}**
- Long-format rows: **{long_rows}**

## Column conventions

The identity columns are:

- `product_id` — GOLDEN product identifier from the workbook.
- `paint_name` — paint name from the workbook.

The wide spectral files use wavelength columns named `wl_400`, `wl_410`, ...,
`wl_700`, where the suffix is the wavelength in nanometres.

The Lab columns in `golden_hb_paints.csv` are named:

- `lab_l_d65_10`
- `lab_a_d65_10`
- `lab_b_d65_10`

In the original workbook these sit under the group header `D65/10°`, meaning
D65 standard illuminant and 10-degree standard observer.

## Suggested Python use

```python
import pandas as pd

reflectance_df = pd.read_csv("golden_hb_reflectance_fraction_wide.csv")
wavelengths = [
    int(column.removeprefix("wl_"))
    for column in reflectance_df.columns
    if column.startswith("wl_")
]
R = reflectance_df[[f"wl_{{w}}" for w in wavelengths]].to_numpy(dtype=float)

print(wavelengths)
print(R.shape)
```

For most colour-science calculations, start with
`golden_hb_reflectance_fraction_wide.csv`. Use `golden_hb_ks_wide.csv` later for
Kubelka-Munk-style mixture modelling.

## Measurement note

The source workbook describes these as **10 mil drawdowns over white**. This
means the measurements are empirical spectra for that paint layer over a white
backing, not necessarily ideal optically opaque pigment reflectance curves.
More transparent paints can include some influence from the white backing.
"""
    (out_dir / "README_golden_clean_csv.md").write_text(readme, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Convert the GOLDEN Heavy Body spectral .xlsx workbook to clean "
            "CSV files."
        )
    )
    parser.add_argument("xlsx", type=Path, help="Path to the source .xlsx workbook")
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=Path("golden_clean_csv"),
        help="Output directory for generated CSV files",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Worksheet name to read. Defaults to the first worksheet.",
    )
    args = parser.parse_args()

    paints = read_golden_workbook(args.xlsx, sheet_name=args.sheet)
    write_clean_csvs(paints, args.out_dir)
    write_readme(args.out_dir, paint_count=len(paints))

    print(f"Wrote {len(paints)} paints to {args.out_dir}")
    print("Generated:")
    for path in sorted(args.out_dir.iterdir()):
        print(f"  {path.name}")


if __name__ == "__main__":
    main()
